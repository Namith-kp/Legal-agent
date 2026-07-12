import os
import sys
import pytest
import tempfile
import docx
import openpyxl
from unittest.mock import patch

# Ensure workspace root is in sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.abspath(os.path.join(current_dir, "..", ".."))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from src.rag import chunk_document, ingest_law_corpus, ingest_session_upload, retrieve
from src.rag.rag_core import Chunk

@pytest.fixture(autouse=True)
def setup_chroma_env(tmp_path):
    # Set the database directory to a temporary path for each test
    db_dir = tmp_path / "chroma_db"
    os.environ["CHROMA_DB_DIR"] = str(db_dir)
    
    # Override the module-level variable dynamically
    from src.rag import rag_core
    old_dir = rag_core.CHROMA_DB_DIR
    rag_core.CHROMA_DB_DIR = str(db_dir)
    
    yield
    
    rag_core.CHROMA_DB_DIR = old_dir
    if "CHROMA_DB_DIR" in os.environ:
        del os.environ["CHROMA_DB_DIR"]

@pytest.fixture
def sample_pdf_file():
    pdf_bytes = b'''%PDF-1.4
1 0 obj
<< /Type /Catalog /Pages 2 0 R >>
endobj
2 0 obj
<< /Type /Pages /Kids [3 0 R 4 0 R] /Count 2 >>
endobj
3 0 obj
<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 5 0 R >> >> /Contents 6 0 R >>
endobj
4 0 obj
<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 5 0 R >> >> /Contents 7 0 R >>
endobj
5 0 obj
<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>
endobj
6 0 obj
<< /Length 85 >>
stream
BT
/F1 12 Tf
100 700 Td
(This is Page 1. Indian Penal Code, 1860. Section 302 details punishment for murder.) Tj
ET
endstream
endobj
7 0 obj
<< /Length 90 >>
stream
BT
/F1 12 Tf
100 700 Td
(This is Page 2. Companies Act, 2013. Section 135 details corporate social responsibility.) Tj
ET
endstream
endobj
xref
0 8
0000000000 65535 f 
0000000009 00000 n 
0000000058 00000 n 
0000000115 00000 n 
0000000224 00000 n 
0000000333 00000 n 
0000000402 00000 n 
0000000538 00000 n 
trailer
<< /Size 8 /Root 1 0 R >>
startxref
679
%%EOF
'''
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
        f.write(pdf_bytes)
        path = f.name
    yield path
    try:
        os.remove(path)
    except OSError:
        pass

@pytest.fixture
def sample_docx_file():
    doc = docx.Document()
    doc.add_paragraph("This is Page 1 of the document. Constitution of India, 1950. Section 12 defines the State.")
    with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as f:
        path = f.name
    doc.save(path)
    yield path
    try:
        os.remove(path)
    except OSError:
        pass

@pytest.fixture
def sample_xlsx_file():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indian Evidence Act"
    ws.cell(row=1, column=1, value="Indian Evidence Act, 1872")
    ws.cell(row=2, column=1, value="Section 3 describes interpretation-clause.")
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    wb.save(path)
    yield path
    try:
        os.remove(path)
    except OSError:
        pass

def test_chunk_pdf(sample_pdf_file):
    chunks = chunk_document(sample_pdf_file)
    assert len(chunks) == 2
    
    # Page 1 validation
    assert "Indian Penal Code" in chunks[0].text
    assert chunks[0].metadata["page"] == "1"
    assert "Indian Penal Code, 1860" in chunks[0].metadata["citation"]
    assert "Section 302" in chunks[0].metadata["citation"]
    
    # Page 2 validation
    assert "Companies Act" in chunks[1].text
    assert chunks[1].metadata["page"] == "2"
    assert "Companies Act, 2013" in chunks[1].metadata["citation"]
    assert "Section 135" in chunks[1].metadata["citation"]

def test_chunk_docx(sample_docx_file):
    chunks = chunk_document(sample_docx_file)
    assert len(chunks) == 1
    assert "Constitution of India" in chunks[0].text
    assert "Constitution of India, 1950" in chunks[0].metadata["citation"]
    assert "Section 12" in chunks[0].metadata["citation"]

def test_chunk_xlsx(sample_xlsx_file):
    chunks = chunk_document(sample_xlsx_file)
    assert len(chunks) == 1
    assert "Indian Evidence Act" in chunks[0].text
    assert "Indian Evidence Act, 1872" in chunks[0].metadata["citation"]
    assert "Section 3" in chunks[0].metadata["citation"]

def test_ingestion_and_retrieval(sample_pdf_file, sample_docx_file):
    # Setup temporary directories for ingestion
    with tempfile.TemporaryDirectory() as corpus_dir:
        # Copy sample PDF to corpus directory
        corpus_pdf_path = os.path.join(corpus_dir, "penal_code.pdf")
        with open(sample_pdf_file, "rb") as f_in:
            with open(corpus_pdf_path, "wb") as f_out:
                f_out.write(f_in.read())
                
        # 1. Ingest corpus
        ingest_law_corpus(corpus_dir)
        
        # 2. Ingest session upload
        session_id = "test-session-123"
        ingest_session_upload(sample_docx_file, session_id)
        
        # 3. Retrieve from both indexes
        # Search query matching both documents
        results = retrieve("Indian Penal Code Section 302", session_id=session_id, score_floor=0.0)
        
        # Results should contain chunks from both the permanent and the session index
        assert len(results) >= 2
        # Check scores sorted in descending order
        scores = [r["score"] for r in results]
        assert scores == sorted(scores, reverse=True)
        
        # Verify specific contents
        citations = [r["source"] for r in results]
        assert any("Indian Penal Code" in c for c in citations)
        assert any("Constitution of India" in c for c in citations)

def test_retrieve_score_floor(sample_pdf_file):
    with tempfile.TemporaryDirectory() as corpus_dir:
        corpus_pdf_path = os.path.join(corpus_dir, "penal_code.pdf")
        with open(sample_pdf_file, "rb") as f_in:
            with open(corpus_pdf_path, "wb") as f_out:
                f_out.write(f_in.read())
        ingest_law_corpus(corpus_dir)
        
        # High score floor should filter out everything
        results_high = retrieve("random query text", score_floor=0.99)
        assert len(results_high) == 0
        
        # Normal score floor should return matches
        results_low = retrieve("Indian Penal Code", score_floor=0.10)
        assert len(results_low) > 0

def test_retrieve_permission_checks(sample_pdf_file):
    with tempfile.TemporaryDirectory() as corpus_dir:
        corpus_pdf_path = os.path.join(corpus_dir, "penal_code.pdf")
        with open(sample_pdf_file, "rb") as f_in:
            with open(corpus_pdf_path, "wb") as f_out:
                f_out.write(f_in.read())
        ingest_law_corpus(corpus_dir)
        
        # Mock permission denied
        with patch("src.sql.check_permission", return_value=False):
            res = retrieve("Indian Penal Code", user_id="user-blocked")
            assert res == {"error": "permission_denied"}
            
        # Mock permission allowed
        with patch("src.sql.check_permission", return_value=True):
            res = retrieve("Indian Penal Code", user_id="user-allowed")
            assert isinstance(res, list)
            assert len(res) > 0
